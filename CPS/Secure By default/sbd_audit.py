import argparse
import os
import sys
import requests
from akamai.edgegrid import EdgeGridAuth, EdgeRc
from urllib.parse import urlparse, urljoin, parse_qsl, urlencode

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[-] Missing 'openpyxl' library. Please run: pip install openpyxl")
    sys.exit(1)

# =========================================================================
# CONFIGURATION VARIABLES
# =========================================================================
# Hardcode your preferences here so you can run the script via simple execution
# =========================================================================
EDGERC_PATH = "~/.edgerc"
EDGERC_SECTION = "default"
ACCOUNT_SWITCH_KEY = "1-IV63:1-2RBL"  # <--- Applied your Context Key
OUTPUT_FILENAME = "akamai_sbd_hostnames.xlsx"
# =========================================================================


def create_sbd_excel(data_list, output_filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SBD Certificate Audit"
    
    # Ensure default grid lines are visible
    ws.views.sheetView[0].showGridLines = True
    
    # Theme: Professional Forest Green
    header_fill = PatternFill(start_color="1E4620", end_color="1E4620", fill_type="solid") 
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    
    zebra_fill = PatternFill(start_color="F5F9F6", end_color="F5F9F6", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    data_font = Font(name="Segoe UI", size=10, color="000000")
    yes_font = Font(name="Segoe UI", size=10, bold=True, color="1E4620")
    
    thin_border_side = Side(border_style="thin", color="D9D9D9")
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    headers = [
        "Property Hostname (cnameFrom)", 
        "Akamai Property Name",
        "Edge Hostname (CNAME Target)", 
        "Edge Hostname ID",
        "SBD on Staging?",
        "SBD on Production?"
    ]
    
    ws.append(headers)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = cell_border
    
    ws.row_dimensions[1].height = 28
    
    for row_idx, item in enumerate(data_list, 2):
        row_data = [
            item["hostname"],
            item["property_name"],
            item["cname_target"],
            item["edge_hostname_id"],
            "YES" if item["staging_sbd"] else "No",
            "YES" if item["production_sbd"] else "No"
        ]
        ws.append(row_data)
        
        current_fill = zebra_fill if row_idx % 2 == 0 else white_fill
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = current_fill
            cell.border = cell_border
            cell.font = data_font
            
            # Formatting alignment for metrics vs names
            if col_idx in [4, 5, 6]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
            if cell.value == "YES":
                cell.font = yes_font
                
        ws.row_dimensions[row_idx].height = 20

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    ws.auto_filter.ref = f"A1:F{len(data_list) + 1}"
    ws.freeze_panes = "A2"
    
    wb.save(output_filename)
    print(f"[+] Audit exported successfully to spreadsheet: {output_filename}")


def main():
    parser = argparse.ArgumentParser(description="Audit Akamai account for Secure By Default (SBD) certificates exclusively.")
    parser.add_argument("--edgerc", default=EDGERC_PATH, help="Path to the Akamai .edgerc file")
    parser.add_argument("--section", default=EDGERC_SECTION, help="The section inside your .edgerc file")
    parser.add_argument("--account-key", dest="account_switch_key", default=ACCOUNT_SWITCH_KEY, help="Akamai Account Switch Key")
    parser.add_argument("--output", default=OUTPUT_FILENAME, help="Output Excel filename")
    
    args = parser.parse_args()
    edgerc_expanded = os.path.expanduser(args.edgerc)
    
    try:
        edgerc = EdgeRc(edgerc_expanded)
        baseurl = edgerc.get(args.section, 'host')
        if not baseurl.startswith(('http://', 'https://')):
            baseurl = 'https://' + baseurl
    except Exception as e:
        print(f"[-] Error loading edgerc configuration from {edgerc_expanded}: {e}", file=sys.stderr)
        sys.exit(1)
        
    session = requests.Session()
    session.auth = EdgeGridAuth(
        client_token=edgerc.get(args.section, 'client_token'),
        client_secret=edgerc.get(args.section, 'client_secret'),
        access_token=edgerc.get(args.section, 'access_token')
    )
    
    url = urljoin(baseurl, "/papi/v1/hostnames")
    
    sbd_map = {}
    total_checked = 0
    
    print(f"[+] Scanning Akamai property assets via endpoint: {url}")
    if args.account_switch_key:
        print(f"[+] Context Switch Applied: accountSwitchKey={args.account_switch_key}")
    else:
        print("[*] Running under default account context.")
        
    while url:
        url_parts = urlparse(url)
        query_params = dict(parse_qsl(url_parts.query))
        
        if args.account_switch_key:
            query_params['accountSwitchKey'] = args.account_switch_key
            
        clean_url = url_parts._replace(query=urlencode(query_params)).geturl()
        headers = {"Accept": "application/json", "PAPI-Use-Prefixes": "true"}
        
        try:
            response = session.get(clean_url, headers=headers)
            if response.status_code != 200:
                print(f"\n[-] Akamai API Error ({response.status_code}): {response.text}", file=sys.stderr)
                sys.exit(1)
            data = response.json()
        except Exception as e:
            print(f"\n[-] Network request failed: {e}", file=sys.stderr)
            sys.exit(1)
            
        hostnames_container = data.get("hostnames", {})
        items = hostnames_container.get("items", []) if isinstance(hostnames_container, dict) else data.get("items", [])
        
        for item in items:
            total_checked += 1
            
            hostname = item.get("cnameFrom")
            if not hostname:
                continue
                
            staging_type = item.get("stagingCertType")
            prod_type = item.get("productionCertType")
            
            # Maps to "DEFAULT" for Secure by Default configurations
            is_staging_sbd = (staging_type == "DEFAULT")
            is_prod_sbd = (prod_type == "DEFAULT")
            
            if is_staging_sbd or is_prod_sbd:
                if hostname not in sbd_map:
                    sbd_map[hostname] = {
                        "hostname": hostname,
                        "property_name": item.get("propertyName") or "N/A",
                        "cname_target": item.get("productionCnameTo") or item.get("stagingCnameTo") or "N/A",
                        "edge_hostname_id": item.get("productionEdgeHostnameId") or item.get("stagingEdgeHostnameId") or "N/A",
                        "staging_sbd": False,
                        "production_sbd": False
                    }
                
                if is_staging_sbd:
                    sbd_map[hostname]["staging_sbd"] = True
                if is_prod_sbd:
                    sbd_map[hostname]["production_sbd"] = True
            
        next_link = hostnames_container.get("nextLink") if isinstance(hostnames_container, dict) else data.get("nextLink")
        if next_link:
            url = urljoin(baseurl, next_link)
            print(".", end="", flush=True)
        else:
            url = None

    print(f"\n[+] Scan complete. Total pipeline configurations examined: {total_checked}")
    print(f"[+] Total Unique Secure By Default (SBD) hostnames discovered: {len(sbd_map)}")
    
    if sbd_map:
        create_sbd_excel(list(sbd_map.values()), args.output)
    else:
        print("[*] No Secure By Default (SBD) hostnames matched in this account view. Excel file not created.")

if __name__ == "__main__":
    main()